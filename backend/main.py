import os
import sys
import uuid
import threading
import traceback
import json
from datetime import datetime
from typing import Dict, Optional, List, Any
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from dotenv import load_dotenv
load_dotenv(os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), ".env"))
from fastapi import FastAPI, UploadFile, File, HTTPException, Query
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse
from fastapi.staticfiles import StaticFiles
from pydantic import BaseModel, Field

from graph.test_analysis_workflow import app as langgraph_app, run_task_analysis
from node.node_list import WordDocumentParser
from db import db_manager

UPLOAD_DIR = os.path.join(os.path.dirname(__file__), "..", "uploads")
os.makedirs(UPLOAD_DIR, exist_ok=True)

app = FastAPI(title="测试点分析系统", version="3.0.0")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

STATIC_DIR = os.path.join(os.path.dirname(__file__), "static")
os.makedirs(STATIC_DIR, exist_ok=True)

app.mount("/static", StaticFiles(directory=STATIC_DIR), name="static")


@app.get("/")
def serve_index():
    frontend_path = os.path.join(STATIC_DIR, "frontend", "index.html")
    if os.path.exists(frontend_path):
        return FileResponse(frontend_path)
    return FileResponse(os.path.join(STATIC_DIR, "index.html"))


# 内存中仅保存运行中的任务进度，最终结果从数据库读取
sessions: Dict[str, dict] = {}


class StartAnalysisInput(BaseModel):
    task_id: Optional[str] = None
    selected_part_ids: List[str] = Field(..., description="选中的需求片段ID列表")


class StartAnalysisBody(BaseModel):
    selected_part_ids: List[str] = Field(..., description="选中的需求片段ID列表")


class RegenerateBody(BaseModel):
    test_point_ids: List[str] = Field(..., description="待重生成的测试点主键 id 列表")
    user_instruction: str = Field(default="", description="用户优化说明")


class PatchTestPointBody(BaseModel):
    is_deleted: Optional[bool] = None


def run_analysis_in_thread(task_id: str, doc_id: str, file_path: str, part_ids: List[str]):
    try:
        # 更新数据库任务状态
        db_manager.update_task_status(task_id, "running")
        
        # 启动 LangGraph 工作流
        # thread_id 使用 task_id
        config = {"configurable": {"thread_id": task_id}}
        
        # 在内存中记录状态，方便前端查询进度
        sessions[task_id] = {
            "status": "analyzing",
            "progress": "准备开始分析...",
            "message": "",
            "task_id": task_id,
            "doc_id": doc_id
        }

        # 执行工作流
        # 我们使用 stream 来获取中间进度更新（可选，这里先简单调用）
        result, _ = run_task_analysis(task_id, doc_id, file_path, part_ids, thread_id=task_id)

        # 任务完成
        db_manager.update_task_status(task_id, "completed")
        
        if task_id in sessions:
            sessions[task_id]["status"] = "completed"
            sessions[task_id]["progress"] = "分析完成"

    except Exception as e:
        error_msg = f"{str(e)}\n{traceback.format_exc()}"
        print(f"Error in analysis thread: {error_msg}")
        db_manager.update_task_status(task_id, "failed", error_message=str(e))
        
        if task_id in sessions:
            sessions[task_id]["status"] = "error"
            sessions[task_id]["progress"] = "分析出错"
            sessions[task_id]["message"] = str(e)


@app.get("/api/health")
def health_check():
    return {"status": "ok"}


@app.post("/api/upload")
async def upload_file(file: UploadFile = File(...)):
    if not file.filename or not file.filename.endswith(".docx"):
        raise HTTPException(status_code=400, detail="仅支持 .docx 文件")

    file_id = str(uuid.uuid4())
    safe_filename = f"{file_id}.docx"
    file_path = os.path.join(UPLOAD_DIR, safe_filename)

    try:
        content = await file.read()
        if not content:
            raise HTTPException(status_code=400, detail="上传文件为空")
        with open(file_path, "wb") as f:
            f.write(content)

        # 解析文档
        parser = WordDocumentParser(file_path)
        parsed_data = parser.parse_section_3()

        if not parsed_data or not parsed_data.sections:
            raise HTTPException(status_code=400, detail="文档解析失败：未找到章节内容")

        # 将解析后的文档保存到数据库，同时获取 part_id 映射
        doc_id, part_id_map = db_manager.save_parsed_document(file.filename, file_path, parsed_data)

        # 生成前端 toc（使用 part_id_map 填充 parts.id）
        toc = []
        for i, sec in enumerate(parsed_data.sections):
            function_types = list(set(p.section_type for p in sec.function_sections))
            part_ids_for_sec = part_id_map[i] if i < len(part_id_map) else []
            parts_data = []
            for j, p in enumerate(sec.function_sections):
                parts_data.append({
                    "id": part_ids_for_sec[j] if j < len(part_ids_for_sec) else "",
                    "type": p.section_type,
                    "content_preview": p.content[:100] + "..." if len(p.content) > 100 else p.content
                })

            toc.append({
                "index": i,
                "title": sec.title,
                "level": sec.level,
                "level_2": sec.metadata.level_2 if sec.metadata else "",
                "level_3": sec.metadata.level_3 if sec.metadata else sec.title,
                "level_4": sec.metadata.level_4 if sec.metadata else "",
                "has_tables": len(sec.tables) > 0,
                "function_types": function_types,
                "parts": parts_data
            })

        return {
            "doc_id": doc_id,
            "file_name": file.filename,
            "status": "uploaded",
            "toc": toc
        }

    except Exception as e:
        print(traceback.format_exc())
        raise HTTPException(
            status_code=500,
            detail=f"文档解析失败: {str(e)}"
        )


@app.get("/api/document/{doc_id}")
def get_document(doc_id: str):
    doc = db_manager.get_document(doc_id)
    if not doc:
        raise HTTPException(status_code=404, detail="文档不存在")
    return doc


@app.get("/api/document-preview/{doc_id}")
def get_document_preview(doc_id: str):
    doc = db_manager.get_document(doc_id)
    if not doc:
        raise HTTPException(status_code=404, detail="文档不存在")

    sections_content = []
    for sec in doc.get("sections", []):
        item = {
            "title": sec["title"],
            "level": sec["level"],
            "function_sections": [],
            "tables": []
        }
        for fs in sec.get("function_sections", []):
            item["function_sections"].append({
                "section_type": fs["section_type"],
                "content": fs["content"]
            })
        for t in sec.get("tables", []):
            headers = t.get("headers")
            rows = t.get("rows")
            if isinstance(headers, str):
                headers = json.loads(headers)
            if isinstance(rows, str):
                rows = json.loads(rows)
            item["tables"].append({
                "headers": headers or [],
                "rows": rows or [],
                "caption": t.get("caption")
            })
        sections_content.append(item)

    return {"sections_content": sections_content}


@app.post("/api/create-task")
def create_task(doc_id: str = Query(...)):
    """为文档创建一个新的分析任务"""
    task_id = db_manager.create_analysis_task(doc_id, [])
    return {"task_id": task_id}


def _start_analysis_impl(task_id: str, part_ids: List[str]) -> Dict[str, Any]:
    if not part_ids:
        raise HTTPException(status_code=400, detail="请选择至少一个需求片段进行分析")

    first_part = db_manager.get_function_part(part_ids[0])
    if not first_part:
        raise HTTPException(status_code=404, detail="需求片段不存在")

    doc_id = str(first_part["doc_id"])
    file_path = first_part["file_path"]

    task_row = db_manager.get_task(task_id)
    if not task_row:
        raise HTTPException(status_code=404, detail="任务不存在，请先创建任务")

    if str(task_row.get("document_id")) != doc_id:
        raise HTTPException(status_code=400, detail="选中片段与任务的文档不一致")

    if not db_manager.update_task_for_start_analysis(task_id, part_ids):
        raise HTTPException(status_code=400, detail="无法更新任务为运行中（任务不存在？）")

    thread = threading.Thread(
        target=run_analysis_in_thread,
        args=(task_id, doc_id, file_path, part_ids),
        daemon=True,
    )
    thread.start()
    return {"task_id": task_id, "status": "started"}


@app.post("/api/start-analysis")
def start_analysis(input_data: StartAnalysisInput):
    """兼容旧前端：可不带 task_id（将自动建任务）或带 task_id。"""
    part_ids = input_data.selected_part_ids
    if not part_ids:
        raise HTTPException(status_code=400, detail="请选择至少一个需求片段进行分析")

    first_part = db_manager.get_function_part(part_ids[0])
    if not first_part:
        raise HTTPException(status_code=404, detail="需求片段不存在")

    doc_id = str(first_part["doc_id"])
    file_path = first_part["file_path"]

    task_id = input_data.task_id
    if not task_id:
        task_id = db_manager.create_analysis_task(doc_id, [])
    return _start_analysis_impl(task_id, part_ids)


@app.post("/api/tasks/{task_id}/start-analysis")
def start_analysis_for_task(task_id: str, body: StartAnalysisBody):
    """规格推荐入口：对已创建任务启动首跑分析。"""
    return _start_analysis_impl(task_id, body.selected_part_ids)


@app.get("/api/task-status/{task_id}")
def get_task_status(task_id: str):
    # 优先从内存获取进度信息
    if task_id in sessions:
        return sessions[task_id]
    
    # 从数据库获取
    task = db_manager.get_task(task_id)
    if task:
        return {
            "task_id": task_id,
            "status": task["status"],
            "progress": "已完成" if task["status"] == "completed" else "未在运行",
            "message": task.get("error_message") or ""
        }

    raise HTTPException(status_code=404, detail="任务不存在")


@app.get("/api/task-results/{task_id}")
def get_task_results(task_id: str):
    """获取任务生成的测试点，同时返回 aggregated_analysis 格式兼容前端"""
    results = db_manager.get_analysis_results(task_id)

    fragments_map = {}
    for tp in results:
        key = tp.get("source_section", "unknown")
        if key not in fragments_map:
            fragments_map[key] = {
                "index": len(fragments_map),
                "section_title": key,
                "content": tp.get("source_content", ""),
                "test_points": []
            }
        steps = tp.get("steps")
        exp_results = tp.get("expected_results")
        if isinstance(steps, str):
            steps = json.loads(steps)
        if isinstance(exp_results, str):
            exp_results = json.loads(exp_results)

        fragments_map[key]["test_points"].append({
            "test_point_id": tp.get("test_point_id"),
            "description": tp.get("description"),
            "priority": tp.get("priority"),
            "test_type": tp.get("test_type"),
            "source_type": tp.get("source_type"),
            "steps": steps or [],
            "expected_results": exp_results or [],
        })

    fragments = sorted(fragments_map.values(), key=lambda f: f["index"])

    aggregated_analysis = {
        "fragments": fragments,
        "total_test_points": len(results),
        "total_fragments": len(fragments),
        "coverage_analysis": f"共 {len(fragments)} 个原文片段，{len(results)} 个测试点"
    }

    return {
        "task_id": task_id,
        "test_points": results,
        "aggregated_analysis": aggregated_analysis,
    }


@app.post("/api/stop-analysis/{task_id}")
def stop_analysis(task_id: str):
    if task_id in sessions:
        # 这里简单标记一下，实际 Graph 节点需要检查 state.is_cancelled
        sessions[task_id]["is_cancelled"] = True
        # 同时更新数据库
        db_manager.update_task_status(task_id, "cancelled")
        return {"status": "cancelling"}
    
    return {"status": "not_running"}


# 导出Excel接口
@app.get("/api/export-test-points")
def export_test_points(
    task_id: Optional[str] = Query(None, description="仅导出该任务下的测试点"),
    include_deleted: bool = Query(False),
):
    """导出测试用例为 Excel；默认仅未软删；可指定 task_id。"""
    results = db_manager.get_all_test_points(task_id=task_id, include_deleted=include_deleted)
    
    from io import BytesIO
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter
    
    wb = Workbook()
    ws = wb.active
    ws.title = "测试用例"
    
    headers = [
        "测试点ID", "测试点描述", "优先级", "测试类型", "用例性质",
        "所属交易", "测试用例目录", "测试步骤", "预期结果", "来源类型", "创建时间"
    ]
    
    header_fill = PatternFill(start_color="4a6cf7", end_color="4a6cf7", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    for row_idx, tp in enumerate(results, 2):
        ws.cell(row=row_idx, column=1, value=tp.get("test_point_id", ""))
        ws.cell(row=row_idx, column=2, value=tp.get("description", ""))
        ws.cell(row=row_idx, column=3, value=tp.get("priority", ""))
        ws.cell(row=row_idx, column=4, value=tp.get("test_type", ""))
        ws.cell(row=row_idx, column=5, value=tp.get("case_nature", ""))
        ws.cell(row=row_idx, column=6, value=tp.get("transaction_name", ""))
        ws.cell(row=row_idx, column=7, value=tp.get("test_case_path", ""))
        
        steps = tp.get("steps", [])
        if isinstance(steps, str):
            import json
            try:
                steps = json.loads(steps)
            except:
                steps = [steps]
        ws.cell(row=row_idx, column=8, value="\n".join(steps))
        
        expected = tp.get("expected_results", [])
        if isinstance(expected, str):
            import json
            try:
                expected = json.loads(expected)
            except:
                expected = [expected]
        ws.cell(row=row_idx, column=9, value="\n".join(expected))
        
        ws.cell(row=row_idx, column=10, value=tp.get("source_type", ""))
        ws.cell(row=row_idx, column=11, value=str(tp.get("created_at", "")))
    
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 20
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    from fastapi.responses import StreamingResponse
    return StreamingResponse(output, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            headers={"Content-Disposition": "attachment; filename=test_points.xlsx"})


# ==================== 任务管理 API ====================

@app.get("/api/tasks")
def get_all_tasks():
    """获取所有任务列表"""
    tasks = db_manager.get_all_tasks()
    result = []
    for task in tasks:
        test_point_count = db_manager.get_test_point_count_by_task(task['id'])
        result.append({
            **task,
            'test_point_count': test_point_count,
        })
    return {"tasks": result}


@app.get("/api/tasks/{task_id}")
def get_task_detail(task_id: str):
    """获取单个任务详情"""
    task = db_manager.get_task(task_id)
    if not task:
        raise HTTPException(status_code=404, detail="任务不存在")
    
    test_point_count = db_manager.get_test_point_count_by_task(task_id)
    
    return {
        **task,
        'test_point_count': test_point_count,
    }


def _run_regen_in_thread(job_id: str) -> None:
    try:
        from services.regeneration_service import run_regeneration_job

        run_regeneration_job(job_id)
    except Exception as e:
        print(traceback.format_exc())
        rjr = getattr(db_manager, "regeneration_jobs_repo", None)
        if rjr is not None:
            try:
                rjr.mark_failed(job_id, str(e))
            except Exception:
                pass


@app.post("/api/tasks/{task_id}/regenerate-test-points")
def regenerate_test_points(task_id: str, body: RegenerateBody):
    rjr = getattr(db_manager, "regeneration_jobs_repo", None)
    if rjr is None:
        raise HTTPException(
            status_code=501,
            detail="当前存储后端不支持重生成（请使用 PostgreSQL 并执行 schema/migrations）",
        )
    if not body.test_point_ids:
        raise HTTPException(status_code=400, detail="请至少选择一个测试点")
    task = db_manager.get_task(task_id)
    if not task:
        raise HTTPException(status_code=404, detail="任务不存在")
    payload = {
        "test_point_ids": body.test_point_ids,
        "user_instruction": body.user_instruction or "",
    }
    job_id = rjr.create_job(task_id, payload)
    threading.Thread(target=_run_regen_in_thread, args=(job_id,), daemon=True).start()
    return {"job_id": job_id}


@app.get("/api/regeneration-jobs/{job_id}")
def get_regeneration_job(job_id: str):
    rjr = getattr(db_manager, "regeneration_jobs_repo", None)
    if rjr is None:
        raise HTTPException(status_code=501, detail="当前存储后端不支持")
    job = rjr.get_job(job_id)
    if not job:
        raise HTTPException(status_code=404, detail="job 不存在")
    return job


@app.get("/api/tasks/{task_id}/test-points")
def list_task_test_points(
    task_id: str,
    limit: int = Query(100, ge=1, le=500),
    offset: int = Query(0, ge=0),
):
    tpr = getattr(db_manager, "test_points_repo", None)
    if tpr is not None:
        return {"test_points": tpr.list_by_task(task_id, limit=limit, offset=offset)}
    rows = db_manager.get_analysis_results(task_id)
    return {"test_points": rows[offset : offset + limit]}


@app.patch("/api/test-points/{test_point_id}")
def patch_test_point(
    test_point_id: str,
    body: PatchTestPointBody,
    task_id: str = Query(..., description="所属任务 id"),
):
    if body.is_deleted is not True:
        raise HTTPException(status_code=400, detail="仅支持 is_deleted=true 软删")
    tpr = getattr(db_manager, "test_points_repo", None)
    if tpr is not None:
        n = tpr.soft_delete(test_point_id, task_id)
    elif hasattr(db_manager, "_get_conn"):
        conn = db_manager._get_conn()
        try:
            cur = conn.execute(
                "UPDATE test_points SET is_deleted = 1, updated_at = ? WHERE id = ? AND task_id = ? AND is_deleted = 0",
                (db_manager._now(), test_point_id, task_id),
            )
            n = cur.rowcount
            conn.commit()
        finally:
            conn.close()
    else:
        raise HTTPException(status_code=501, detail="不支持该操作")
    if not n:
        raise HTTPException(status_code=404, detail="测试点不存在或已删除")
    return {"status": "ok"}


@app.delete("/api/tasks/{task_id}")
def delete_task(task_id: str):
    """删除任务及其相关数据（依赖外键 CASCADE）。"""
    try:
        db_manager.delete_task_cascade(task_id)
        return {"status": "success", "message": "任务已删除"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"删除失败: {str(e)}")
